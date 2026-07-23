#!/usr/bin/env python3
"""addi — fill the UK DRI FAIR-metadata Excel workbook for an ADDI / AD Workbench
data-portal submission.

Unlike the other skills in this collection (which are stdlib-only), `addi` reads and
writes a richly-formatted `.xlsx` template, so it depends on **openpyxl** (reading the
template's structure, controlled vocabularies and validations) and **pandas** (reading
the dictionaries/fields/lookups tables and an optional metadata.tsv). See DESIGN.md.

The template ships with dropdown data-validations (some referencing a hidden
`Catalogue_Data` vocab sheet via the x14 extension), cell colors, per-cell prompts,
threaded comments and a document sensitivity label. openpyxl's writer silently drops
most of these on save, so `fill` does **not** save through openpyxl. Instead it edits
the template *surgically*: it copies every part of the original workbook verbatim and
rewrites only the `<sheetData>` of the sheets it changes, using inline strings so the
shared-strings table is left untouched. Everything else — dropdowns, colors, comments,
the hidden vocab sheet — is preserved byte-for-byte.

Subcommands:
  info      describe the template (sheets, mandatory fields, dropdown vocabularies)
  fill      write user-supplied values into the template -> a submission workbook
  validate  check a filled workbook (or the inputs) against the template's rules

Every command accepts --template PATH to target a different template file (default:
the shipped V1.2 workbook next to this script).
"""

import argparse
import json
import os
import re
import sys
import warnings
import zipfile

# ---------------------------------------------------------------------------
# Third-party dependencies (see DESIGN.md — the one documented exception to the
# collection's stdlib-only rule).
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("addi requires 'openpyxl' (pip install openpyxl). See DESIGN.md.")
try:
    import pandas as pd
except ImportError:  # pragma: no cover
    sys.exit("addi requires 'pandas' (pip install pandas). See DESIGN.md.")

# openpyxl warns that it will drop the x14 data-validation extension on load. We never
# save through openpyxl (see module docstring), so the warning is noise here.
warnings.filterwarnings("ignore", message=".*Data Validation extension.*")

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(
    HERE, "..", "templates",
    "template_UK_DRI_FAIR_Metadata_w_ExtendedCatalogue_V1.2..xlsx",
)

# Sheet names in the shipped template (note the trailing space in "fields ").
SHEET_CATALOGUE = "catalogue"
SHEET_EXTENDED = "extendedcatalogue"
SHEET_SETTINGS = "settings"
SHEET_WORKSPACE = "workspace_settings"
SHEET_DICTIONARIES = "dictionaries"
SHEET_FIELDS = "fields "
SHEET_LOOKUPS = "lookups"
SHEET_VOCAB = "Catalogue_Data"

# Mandatory catalogue keys per the template README, note (4).
CATALOGUE_MANDATORY = ["title", "description", "publisher_name"]

# Always required from user-provided input — a deliberate addi-specific addition
# beyond the README's mandatory set (see DESIGN.md). NEVER guessed or defaulted.
USER_ATTRIBUTED = [
    ("catalogue", "creator"),
    ("catalogue", "contactPoint"),
    ("workspace_settings", "dataset owners"),
]

# settings.visibility allowed values per README note (1).
VISIBILITY_ALLOWED = ["private", "internal"]

# fields.type allowed values per README note (11).
FIELD_TYPES = ["boolean", "date", "datetime", "decimal", "integer", "text", "time"]

# code / name syntax rules per README notes (3), (8), (9), (10).
RE_CODE = re.compile(r"^[A-Za-z0-9_]+$")            # letters, numbers, underscore
RE_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")   # + cannot start with a number

VALUE_COLS = ["D", "E", "F", "G", "H", "I"]         # extendedcatalogue Value1..Value6


# ===========================================================================
# Output hygiene (xlsx cell model — see DESIGN.md "Clean output fields")
# ===========================================================================
_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")  # legal in OOXML: \t \n \r


def clean_text(value):
    """Reduce a value to a safe xlsx text token.

    Newlines are legal inside a cell and are preserved; the control characters the
    OOXML format forbids are stripped; and a value beginning with =, +, - or @ is
    prefixed with an apostrophe to neutralize spreadsheet formula injection.
    """
    s = "" if value is None else str(value)
    s = _CONTROL.sub("", s)
    if s and s[0] in "=+-@":
        s = "'" + s
    return s


def xml_escape(s):
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
             .replace('"', "&quot;"))


# ===========================================================================
# Column-letter helpers
# ===========================================================================
def col_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def split_ref(ref):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    return m.group(1), int(m.group(2))


# ===========================================================================
# Reading the template with openpyxl
# ===========================================================================
class Template:
    """Reads structure, mandatory fields and controlled vocabularies from a template
    workbook. Read-only; writing is done surgically by fill()."""

    def __init__(self, path):
        if not os.path.exists(path):
            sys.exit(f"template not found: {path}")
        self.path = path
        self.wb = openpyxl.load_workbook(path, data_only=True)
        for name in (SHEET_CATALOGUE, SHEET_EXTENDED, SHEET_SETTINGS,
                     SHEET_WORKSPACE, SHEET_DICTIONARIES, SHEET_FIELDS,
                     SHEET_LOOKUPS, SHEET_VOCAB):
            if name not in self.wb.sheetnames:
                sys.exit(f"template is missing the '{name}' sheet — is this the "
                         f"expected UK DRI FAIR-metadata workbook?")
        self._raw_ext = self._read_raw_sheet_xml(SHEET_EXTENDED)
        self.ext_fields = self._read_extended_fields()

    # -- raw XML access (for parsing data-validations that openpyxl drops) ----
    def _sheet_filename(self, sheet_title):
        """Map a sheet title to its xl/worksheets/sheetN.xml path via the rels."""
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(self.path) as z:
            wbxml = z.read("xl/workbook.xml").decode("utf-8")
            relsxml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        # name -> r:id
        ns_rid = dict(re.findall(r'<sheet[^>]*name="([^"]*)"[^>]*r:id="(rId\d+)"',
                                 wbxml))
        # r:id -> target
        rid_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/[^"]*)"',
                                     relsxml))
        rid = ns_rid.get(sheet_title)
        target = rid_target.get(rid)
        if not target:
            sys.exit(f"could not resolve worksheet file for sheet '{sheet_title}'")
        return "xl/" + target

    def _read_raw_sheet_xml(self, sheet_title):
        fn = self._sheet_filename(sheet_title)
        with zipfile.ZipFile(self.path) as z:
            return z.read(fn).decode("utf-8")

    # -- extendedcatalogue field discovery -----------------------------------
    def _read_extended_fields(self):
        """Return ordered list of dicts describing rows 7..20 of extendedcatalogue."""
        ws = self.wb[SHEET_EXTENDED]
        validations = self._parse_ext_validations()
        fields = []
        for row in range(7, 21):
            raw_name = ws.cell(row=row, column=1).value
            if raw_name is None or str(raw_name).strip() == "":
                continue
            raw_name = str(raw_name).strip()
            mandatory = raw_name.startswith("*")
            label = raw_name.lstrip("*").strip()
            ftype_txt = (ws.cell(row=row, column=2).value or "").strip().lower()
            allowed = validations.get(row)  # list of allowed values, or None
            multi = ("multi-select" in ftype_txt) or ("multi-select" in label.lower())
            if "integer" in ftype_txt:
                kind = "integer"
            elif "toggle" in ftype_txt:
                kind = "toggle"
            elif "url" in ftype_txt:
                kind = "url"
            elif allowed is not None:
                kind = "list"
            else:
                kind = "text"
            fields.append({
                "row": row,
                "label": label,
                "norm": normalize_key(label),
                "mandatory": mandatory,
                "type_text": ftype_txt,
                "kind": kind,
                "multi": multi and kind not in ("integer", "toggle"),
                "allowed": allowed,
            })
        return fields

    def _parse_ext_validations(self):
        """Map extendedcatalogue row-number -> list of allowed dropdown values.

        Handles both the standard inline list (e.g. the Yes/No toggle) and the x14
        extension validations that reference a Catalogue_Data column range.
        """
        out = {}
        xml = self._raw_ext
        # Standard <dataValidation type="list" sqref="D12"><formula1>"Yes, No"</...>
        for m in re.finditer(
                r'<dataValidation\b[^>]*\btype="list"[^>]*\bsqref="([^"]+)"[^>]*>'
                r'\s*<formula1>(.*?)</formula1>', xml, re.S):
            sqref, formula = m.group(1), m.group(2).strip()
            row = self._sqref_start_row(sqref)
            lit = formula.strip().strip('"')
            values = [v.strip() for v in lit.split(",") if v.strip()]
            if row and values:
                out[row] = values
        # x14 validations: <xm:f>Catalogue_Data!$G$2:$G$28</xm:f> ... <xm:sqref>D20:I20</xm:sqref>
        for m in re.finditer(
                r'<x14:dataValidation\b.*?<xm:f>(.*?)</xm:f>.*?'
                r'<xm:sqref>(.*?)</xm:sqref>.*?</x14:dataValidation>', xml, re.S):
            formula, sqref = m.group(1).strip(), m.group(2).strip()
            row = self._sqref_start_row(sqref)
            values = self._resolve_range(formula)
            if row and values:
                out[row] = values
        return out

    @staticmethod
    def _sqref_start_row(sqref):
        first = sqref.split()[0].split(":")[0]
        m = re.match(r"[A-Z]+(\d+)", first)
        return int(m.group(1)) if m else None

    def _resolve_range(self, formula):
        """Resolve 'Catalogue_Data!$G$2:$G$28' -> list of cell values from the sheet."""
        m = re.match(r"([^!]+)!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", formula)
        if not m:
            return None
        sheet, c1, r1, c2, r2 = m.groups()
        sheet = sheet.strip("'")
        if sheet not in self.wb.sheetnames:
            return None
        ws = self.wb[sheet]
        col = col_to_num(c1)
        values = []
        for r in range(int(r1), int(r2) + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip():
                values.append(str(v).strip())
        return values

    # -- vocab dump for `info` ------------------------------------------------
    def vocab_columns(self):
        """Return {header: [values]} for every column of the hidden vocab sheet."""
        ws = self.wb[SHEET_VOCAB]
        out = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not header:
                continue
            vals = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is not None and str(v).strip():
                    vals.append(str(v).strip())
            out[str(header).strip()] = vals
        return out

    def catalogue_keys(self):
        ws = self.wb[SHEET_CATALOGUE]
        return [str(ws.cell(row=r, column=1).value).strip()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value]

    def settings_keys(self):
        ws = self.wb[SHEET_SETTINGS]
        return [str(ws.cell(row=r, column=1).value).strip()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value]

    def workspace_keys(self):
        ws = self.wb[SHEET_WORKSPACE]
        # keys live in column B (column A is empty for data rows)
        return [str(ws.cell(row=r, column=2).value).strip()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=2).value]


def normalize_key(label):
    """Normalize a field label for matching: drop leading '*', a trailing
    parenthetical like '(Multi-select)', collapse whitespace, lowercase."""
    s = str(label).lstrip("*").strip()
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s)
    return re.sub(r"\s+", " ", s).strip().lower()


# ===========================================================================
# Surgical worksheet XML editing
# ===========================================================================
class SheetXML:
    """Edit a single worksheet's <sheetData> in place, preserving everything else
    (styles, dropdowns, extLst, comments references)."""

    def __init__(self, xml):
        self.xml = xml
        m = re.search(r"(<sheetData\b[^>]*>)(.*?)(</sheetData>)", xml, re.S)
        if not m:
            raise ValueError("no <sheetData> in worksheet XML")
        self._pre = xml[:m.start()]
        self._open = m.group(1)
        self._body = m.group(2)
        self._close = m.group(3)
        self._post = xml[m.end():]
        self.rows = self._parse_rows(self._body)  # {row_num: {"attrs":str, "cells":{ref:cellxml}}}

    @staticmethod
    def _parse_rows(body):
        rows = {}
        for rm in re.finditer(r"<row\b([^>]*?)(?:/>|>(.*?)</row>)", body, re.S):
            attrs, inner = rm.group(1), rm.group(2) or ""
            num = int(re.search(r'\br="(\d+)"', attrs).group(1))
            cells = {}
            for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)".*?(?:/>|>.*?</c>)',
                                  inner, re.S):
                cells[cm.group(1)] = cm.group(0)
            rows[num] = {"attrs": attrs, "cells": cells}
        return rows

    @staticmethod
    def _style_of(cell_xml):
        if cell_xml is None:
            return None
        m = re.search(r'\bs="(\d+)"', cell_xml)
        return m.group(1) if m else None

    @staticmethod
    def build_cell(ref, value, style=None, kind="auto"):
        """Build a <c> element. kind: 'str', 'num', 'bool', 'empty', or 'auto'."""
        s_attr = f' s="{style}"' if style is not None else ""
        if kind == "empty" or value is None or (kind == "auto" and value == ""):
            return f'<c r="{ref}"{s_attr}/>'
        if kind == "bool":
            b = "1" if bool_from(value) else "0"
            return f'<c r="{ref}"{s_attr} t="b"><v>{b}</v></c>'
        if kind == "num" or (kind == "auto" and isinstance(value, (int, float))
                             and not isinstance(value, bool)):
            num = value
            if isinstance(num, float) and num.is_integer():
                num = int(num)
            return f'<c r="{ref}"{s_attr}><v>{num}</v></c>'
        text = xml_escape(clean_text(value))
        return (f'<c r="{ref}"{s_attr} t="inlineStr"><is>'
                f'<t xml:space="preserve">{text}</t></is></c>')

    def set_cell(self, ref, value, kind="auto", style="keep", default_style=None):
        """Set/replace/insert a cell. style='keep' preserves the existing cell's
        style (falling back to default_style if the cell is absent)."""
        col, rownum = split_ref(ref)
        if rownum not in self.rows:
            self.rows[rownum] = {"attrs": f' r="{rownum}"', "cells": {}}
        row = self.rows[rownum]
        existing = row["cells"].get(ref)
        st = self._style_of(existing) if style == "keep" else style
        if st is None and default_style is not None:
            st = default_style
        row["cells"][ref] = self.build_cell(ref, value, style=st, kind=kind)

    def replace_data_rows(self, header_row, new_rows):
        """Keep rows up to and including header_row; replace all rows below with
        new_rows = list of {ref: <c-xml>} dicts (already built)."""
        self.rows = {n: r for n, r in self.rows.items() if n <= header_row}
        for i, cells in enumerate(new_rows):
            self.rows[header_row + 1 + i] = {
                "attrs": f' r="{header_row + 1 + i}" spans="1:9"',
                "cells": cells,
            }

    def drop_hyperlinks(self, refs):
        """Remove <hyperlink> entries for the given cell refs (they live after
        </sheetData>), so a cell we overwrote or cleared does not keep a stale link
        (e.g. the template's mailto: on contactPoint). Orphaned external relationships
        in the sheet .rels are harmless and left in place."""
        refset = set(refs)
        if "<hyperlinks" not in self._post:
            return

        def repl(m):
            kept = [hl for hl in re.findall(r"<hyperlink\b.*?(?:/>|</hyperlink>)",
                                            m.group(1), re.S)
                    if re.search(r'\bref="([^"]+)"', hl).group(1) not in refset]
            return f"<hyperlinks>{''.join(kept)}</hyperlinks>" if kept else ""

        self._post = re.sub(r"<hyperlinks>(.*?)</hyperlinks>", repl, self._post, flags=re.S)

    def serialize(self):
        parts = []
        for num in sorted(self.rows):
            row = self.rows[num]
            cells = row["cells"]
            ordered = sorted(cells, key=lambda ref: col_to_num(split_ref(ref)[0]))
            inner = "".join(cells[ref] for ref in ordered)
            if inner:
                parts.append(f"<row{row['attrs']}>{inner}</row>")
            else:
                parts.append(f"<row{row['attrs']}/>")
        return self._pre + self._open + "".join(parts) + self._close + self._post


def bool_from(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes", "y", "t")


# ===========================================================================
# Config loading
# ===========================================================================
def load_config(path):
    if path is None:
        return {}
    if not os.path.exists(path):
        sys.exit(f"config not found: {path}")
    text = open(path, encoding="utf-8").read()
    if path.lower().endswith((".yml", ".yaml")):
        try:
            import yaml
        except ImportError:
            sys.exit("PyYAML is not installed; convert the config to JSON or "
                     "'pip install pyyaml'.")
        return yaml.safe_load(text) or {}
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        sys.exit(f"could not parse JSON config {path}: {e}")


def read_table(path):
    """Read a dictionaries/fields/lookups table (TSV or CSV) into a DataFrame."""
    if not os.path.exists(path):
        sys.exit(f"table not found: {path}")
    sep = "\t" if path.lower().endswith((".tsv", ".txt")) else ","
    df = pd.read_csv(path, sep=sep, dtype=str, keep_default_na=False)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ===========================================================================
# metadata.tsv seeding
# ===========================================================================
def infer_field_type(series):
    vals = [v for v in series.tolist() if str(v).strip() not in ("", "NA")]
    if not vals:
        return "text"
    def all_match(fn):
        for v in vals:
            try:
                fn(v)
            except (ValueError, TypeError):
                return False
        return True
    if all_match(lambda v: int(str(v))):
        return "integer"
    if all_match(lambda v: float(str(v))):
        return "decimal"
    if all(re.match(r"^\d{4}-\d{2}-\d{2}$", str(v)) for v in vals):
        return "date"
    return "text"


def sanitize_name(col):
    name = re.sub(r"[^A-Za-z0-9_]", "_", str(col).strip())
    if not name or name[0].isdigit():
        name = "_" + name
    return name


def seed_from_metadata(path, dict_code, warn):
    """Return (dictionaries_df, fields_df, extended_seed) derived from a metadata.tsv."""
    df = read_table(path)
    if not RE_CODE.match(dict_code) or dict_code[0].isdigit():
        sys.exit(f"--metadata-dict-code '{dict_code}' must be letters/numbers/"
                 f"underscore and not start with a number.")
    dictionaries = pd.DataFrame(
        [{"code": dict_code, "name": dict_code.replace("_", " ").title(),
          "description": f"Seeded from {os.path.basename(path)}"}])
    frows = []
    for col in df.columns:
        name = sanitize_name(col)
        if name != col:
            warn(f"metadata column '{col}' -> field name '{name}'")
        frows.append({
            "dictionary_code": dict_code, "name": name, "label": str(col),
            "type": infer_field_type(df[col]), "constraints": "",
            "description": "", "uri": "",
            "entity": "true" if name.lower() in ("id", "sample", "sample_id") else "false",
            "cohort_filter": "false",
        })
    fields = pd.DataFrame(frows)

    # Conservative descriptive seeding (guesses — warned, overridden by --config).
    extended_seed = {}
    id_col = next((c for c in df.columns if c.lower() in ("sample", "sample_id", "id")),
                  None)
    n = int(df[id_col].nunique()) if id_col else int(len(df))
    extended_seed["Number of Biosamples"] = n
    warn(f"seeded 'Number of Biosamples' = {n} (from metadata.tsv; review).")
    return dictionaries, fields, extended_seed


# ===========================================================================
# Validation
# ===========================================================================
def validate_inputs(tpl, cfg, dictionaries, fields, lookups):
    """Return a list of (severity, message). severity in {'ERROR','WARN'}."""
    issues = []
    cat = cfg.get("catalogue", {}) or {}
    ws = cfg.get("workspace_settings", {}) or {}
    settings = cfg.get("settings", {}) or {}
    extended = cfg.get("extendedcatalogue", {}) or {}

    def is_empty(v):
        return v is None or str(v).strip() == ""

    # catalogue mandatory (README note 4)
    for key in CATALOGUE_MANDATORY:
        if is_empty(cat.get(key)):
            issues.append(("ERROR", f"catalogue.{key} is mandatory (README note 4)."))

    # always-required user-attributed fields (addi exception — see DESIGN.md)
    for sheet, key in USER_ATTRIBUTED:
        src = cat if sheet == "catalogue" else ws
        if is_empty(src.get(key)):
            issues.append(("ERROR",
                           f"{sheet}.'{key}' must be provided by the user "
                           f"(never inferred; see DESIGN.md)."))

    # settings.visibility (README note 1)
    vis = settings.get("visibility")
    if vis is not None and str(vis).strip() and str(vis).strip() not in VISIBILITY_ALLOWED:
        issues.append(("ERROR", f"settings.visibility '{vis}' must be one of "
                                f"{VISIBILITY_ALLOWED} (README note 1)."))

    # extendedcatalogue: mandatory + dropdown membership + multi-select cap
    for f in tpl.ext_fields:
        provided = _lookup_ci(extended, f["label"])
        vals = as_list(provided)
        if f["mandatory"] and not vals:
            issues.append(("ERROR", f"extendedcatalogue '{f['label']}' is mandatory."))
        if f["allowed"] is not None and vals:
            allowed_ci = {a.lower(): a for a in f["allowed"]}
            for v in vals:
                if str(v).strip().lower() not in allowed_ci:
                    issues.append(("ERROR", f"extendedcatalogue '{f['label']}' value "
                                            f"'{v}' is not in the allowed dropdown list."))
        if f["multi"] and len(vals) > len(VALUE_COLS):
            issues.append(("ERROR", f"extendedcatalogue '{f['label']}' has {len(vals)} "
                                    f"values but only {len(VALUE_COLS)} are allowed."))
        if not f["multi"] and len(vals) > 1:
            issues.append(("ERROR", f"extendedcatalogue '{f['label']}' is single-value "
                                    f"but {len(vals)} values were given."))
        if f["kind"] == "integer" and vals:
            try:
                int(str(vals[0]))
            except ValueError:
                issues.append(("ERROR", f"extendedcatalogue '{f['label']}' must be an "
                                        f"integer, got '{vals[0]}'."))

    issues += validate_schema(dictionaries, fields, lookups)
    return issues


def validate_schema(dictionaries, fields, lookups):
    issues = []
    dict_codes = set()
    if dictionaries is not None:
        for _, r in dictionaries.iterrows():
            code = str(r.get("code", "")).strip()
            if not code:
                continue
            dict_codes.add(code)
            if not RE_CODE.match(code) or code[0].isdigit():
                issues.append(("ERROR", f"dictionaries.code '{code}' must be letters/"
                                        f"numbers/underscore and not start with a number "
                                        f"(README notes 3, 8)."))
    lookup_names = set()
    if lookups is not None:
        for _, r in lookups.iterrows():
            lk = str(r.get("lookup", "")).strip()
            if lk:
                lookup_names.add(lk)
    if fields is not None:
        for _, r in fields.iterrows():
            name = str(r.get("name", "")).strip()
            dc = str(r.get("dictionary_code", "")).strip()
            ftype = str(r.get("type", "")).strip().lower()
            constraints = str(r.get("constraints", "")).strip()
            if name and not RE_NAME.match(name):
                issues.append(("ERROR", f"fields.name '{name}' must be letters/numbers/"
                                        f"underscore and not start with a number "
                                        f"(README note 10)."))
            if dc and dict_codes and dc not in dict_codes:
                issues.append(("ERROR", f"fields.dictionary_code '{dc}' has no matching "
                                        f"row in dictionaries (README note 9)."))
            if ftype and ftype not in FIELD_TYPES:
                issues.append(("ERROR", f"fields.type '{ftype}' must be one of "
                                        f"{FIELD_TYPES} (README note 11)."))
            if ftype == "boolean" and constraints:
                issues.append(("ERROR", f"fields '{name}' is boolean and cannot have "
                                        f"constraints (README note 12)."))
            if constraints and lookup_names and constraints not in lookup_names:
                issues.append(("WARN", f"fields.constraints '{constraints}' has no "
                                       f"matching lookup in the lookups sheet "
                                       f"(README note 13)."))
    return issues


def _lookup_ci(d, key):
    """Case/normalize-insensitive lookup into a dict by field label."""
    if key in d:
        return d[key]
    nk = normalize_key(key)
    for k, v in d.items():
        if normalize_key(k) == nk:
            return v
    return None


def as_list(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [v for v in value if str(v).strip() != ""]
    s = str(value).strip()
    return [s] if s else []


# ===========================================================================
# Subcommands
# ===========================================================================
def cmd_info(args):
    tpl = Template(args.template)
    if args.json:
        out = {
            "sheets": tpl.wb.sheetnames,
            "catalogue_keys": tpl.catalogue_keys(),
            "catalogue_mandatory": CATALOGUE_MANDATORY,
            "user_attributed_required": [f"{s}.{k}" for s, k in USER_ATTRIBUTED],
            "settings_keys": tpl.settings_keys(),
            "settings_visibility_allowed": VISIBILITY_ALLOWED,
            "workspace_keys": tpl.workspace_keys(),
            "extendedcatalogue_fields": [
                {"label": f["label"], "mandatory": f["mandatory"],
                 "kind": f["kind"], "multi": f["multi"], "allowed": f["allowed"]}
                for f in tpl.ext_fields],
            "field_types": FIELD_TYPES,
            "vocabularies": tpl.vocab_columns(),
        }
        print(json.dumps(out, indent=2, ensure_ascii=False))
        return

    print(f"Template: {os.path.basename(tpl.path)}")
    print(f"Sheets:   {', '.join(tpl.wb.sheetnames)}\n")
    print("catalogue keys (mandatory marked *):")
    for k in tpl.catalogue_keys():
        star = " *" if k in CATALOGUE_MANDATORY else ""
        print(f"  {k}{star}")
    print("\nAlways required from the user (addi rule — never inferred):")
    for s, k in USER_ATTRIBUTED:
        print(f"  {s}.{k}")
    print(f"\nsettings.visibility allowed: {', '.join(VISIBILITY_ALLOWED)}")
    print("\nextendedcatalogue fields (mandatory marked *):")
    for f in tpl.ext_fields:
        star = "*" if f["mandatory"] else " "
        tag = f["kind"] + ("[multi]" if f["multi"] else "")
        n = f" — {len(f['allowed'])} allowed values" if f["allowed"] else ""
        print(f"  {star} {f['label']}  ({tag}){n}")
    print(f"\nfields.type allowed: {', '.join(FIELD_TYPES)}")
    print("\nControlled vocabularies (hidden Catalogue_Data sheet):")
    for name, vals in tpl.vocab_columns().items():
        preview = ", ".join(vals[:8]) + (" ..." if len(vals) > 8 else "")
        print(f"  {name} ({len(vals)}): {preview}")
    print("\nUse --json for the full machine-readable dump (all allowed values).")


def read_filled_workbook(tpl, path):
    """Read values back out of a filled workbook into the same (cfg, dictionaries,
    fields, lookups) shape the input path uses, so validate can check either."""
    if not os.path.exists(path):
        sys.exit(f"workbook not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    def kv(sheet, key_col, val_col):
        ws = wb[sheet]
        out = {}
        for r in range(2, ws.max_row + 1):
            k = ws.cell(row=r, column=key_col).value
            if k is None or str(k).strip() == "":
                continue
            out[str(k).strip()] = ws.cell(row=r, column=val_col).value
        return out

    cfg = {
        "catalogue": kv(SHEET_CATALOGUE, 1, 2),
        "settings": kv(SHEET_SETTINGS, 1, 2),
        "workspace_settings": kv(SHEET_WORKSPACE, 2, 3),
        "extendedcatalogue": {},
    }
    ext_ws = wb[SHEET_EXTENDED]
    for f in tpl.ext_fields:
        vals = []
        for col in range(col_to_num(VALUE_COLS[0]), col_to_num(VALUE_COLS[-1]) + 1):
            v = ext_ws.cell(row=f["row"], column=col).value
            if v is not None and str(v).strip() != "":
                vals.append(v)
        if vals:
            cfg["extendedcatalogue"][f["label"]] = vals

    def table(sheet):
        ws = wb[sheet]
        headers = [str(ws.cell(row=1, column=c).value).strip()
                   for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in vals):
                continue
            rows.append({h: ("" if v is None else str(v)) for h, v in zip(headers, vals)})
        return pd.DataFrame(rows) if rows else None

    return cfg, table(SHEET_DICTIONARIES), table(SHEET_FIELDS), table(SHEET_LOOKUPS)


def cmd_validate(args):
    tpl = Template(args.template)
    if args.workbook:
        cfg, dictionaries, fields, lookups = read_filled_workbook(tpl, args.workbook)
    else:
        cfg = load_config(args.config)
        dictionaries = read_table(args.dictionaries) if args.dictionaries else None
        fields = read_table(args.fields) if args.fields else None
        lookups = read_table(args.lookups) if args.lookups else None
    issues = validate_inputs(tpl, cfg, dictionaries, fields, lookups)
    _report(issues)
    if any(sev == "ERROR" for sev, _ in issues):
        sys.exit(1)
    print("OK — inputs satisfy the template's rules.", file=sys.stderr)


def _report(issues):
    for sev, msg in issues:
        print(f"{sev}: {msg}", file=sys.stderr)


def cmd_fill(args):
    tpl = Template(args.template)
    cfg = load_config(args.config)
    warnings_out = []
    warn = lambda m: warnings_out.append(m)

    # --- assemble schema tables, applying precedence: metadata < config-file tables
    dictionaries = fields = lookups = None
    ext_seed = {}
    if args.metadata:
        dictionaries, fields, ext_seed = seed_from_metadata(
            args.metadata, args.metadata_dict_code, warn)
    if args.dictionaries:
        dictionaries = read_table(args.dictionaries)
    if args.fields:
        fields = read_table(args.fields)
    if args.lookups:
        lookups = read_table(args.lookups)

    # --- merge extendedcatalogue: metadata seed (lowest) < config (highest)
    extended = dict(ext_seed)
    for k, v in (cfg.get("extendedcatalogue", {}) or {}).items():
        extended[k] = v
    cfg = dict(cfg)
    cfg["extendedcatalogue"] = extended

    # --- validate before writing (fail fast; user-attributed enforcement lives here)
    issues = validate_inputs(tpl, cfg, dictionaries, fields, lookups)
    _report(issues)
    if any(sev == "ERROR" for sev, _ in issues):
        sys.exit("Refusing to write: fix the ERROR(s) above. "
                 "creator / contactPoint / dataset owners must come from you.")

    # --- surgical write
    edited = build_filled_workbook(tpl, cfg, dictionaries, fields, lookups)
    write_zip(tpl.path, args.out, edited)

    for w in warnings_out:
        print(f"WARN: {w}", file=sys.stderr)
    print(f"Wrote {args.out}", file=sys.stderr)
    print("Review the seeded/guessed values before submitting to AD Workbench.",
          file=sys.stderr)


# ===========================================================================
# Building the filled workbook (returns {sheet_file_path: new_xml})
# ===========================================================================
def build_filled_workbook(tpl, cfg, dictionaries, fields, lookups):
    edited = {}

    # catalogue: key in col A (rows 2..N), value in col B. Provided keys are written;
    # unprovided keys whose template value is an ALL-CAPS placeholder are cleared, while
    # genuine defaults (license, publisher_url, language, contactPoint, …) are kept.
    cat = cfg.get("catalogue", {}) or {}
    fn = tpl._sheet_filename(SHEET_CATALOGUE)
    sx = _load_sheet(tpl, fn)
    ws = tpl.wb[SHEET_CATALOGUE]
    cat_modified = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        if not key:
            continue
        key = str(key).strip()
        if key in cat:
            val = cat[key]
            kind = "empty" if str(val).strip() == "" else "str"
            sx.set_cell(f"B{r}", val, kind=kind, default_style="4")
            cat_modified.append(f"B{r}")
        elif is_caps_placeholder(ws.cell(row=r, column=2).value):
            sx.set_cell(f"B{r}", "", kind="empty")
            cat_modified.append(f"B{r}")
    sx.drop_hyperlinks(cat_modified)  # e.g. the stale mailto: on an overwritten contactPoint
    edited[fn] = sx.serialize()

    # settings: key in col A, value in col B (boolean rows use t="b")
    settings = cfg.get("settings", {}) or {}
    if settings:
        fn = tpl._sheet_filename(SHEET_SETTINGS)
        sx = _load_sheet(tpl, fn)
        ws = tpl.wb[SHEET_SETTINGS]
        keymap = {str(ws.cell(row=r, column=1).value).strip(): r
                  for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}
        bool_keys = {"allow_private_cohorts", "allow_internal_cohorts",
                     "expose_cohort_counts", "expose_cohort_visualisations",
                     "allow_dataset_update_subscriptions", "allow_clear",
                     "allow_pseudonymised"}
        for key, val in settings.items():
            if key not in keymap:
                continue
            kind = "bool" if key in bool_keys else "str"
            sx.set_cell(f"B{keymap[key]}", val, kind=kind)
        edited[fn] = sx.serialize()

    # workspace_settings: key in col B, value in col C
    wsp = cfg.get("workspace_settings", {}) or {}
    if wsp:
        fn = tpl._sheet_filename(SHEET_WORKSPACE)
        sx = _load_sheet(tpl, fn)
        ws = tpl.wb[SHEET_WORKSPACE]
        keymap = {str(ws.cell(row=r, column=2).value).strip(): r
                  for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=2).value}
        for key, val in wsp.items():
            if key not in keymap:
                continue
            sx.set_cell(f"C{keymap[key]}", val, kind="str", default_style=None)
        edited[fn] = sx.serialize()

    # extendedcatalogue: values in D..I per field row. Every field row is processed so
    # that any value cell the user did not provide is cleared of its shipped example.
    extended = cfg.get("extendedcatalogue", {}) or {}
    fn = tpl._sheet_filename(SHEET_EXTENDED)
    sx = _load_sheet(tpl, fn)
    ext_modified = []
    for f in tpl.ext_fields:
        provided = _lookup_ci(extended, f["label"])
        vals = as_list(provided) if provided is not None else []
        ncols = len(VALUE_COLS) if f["multi"] else 1
        for i in range(ncols):
            ref = f"{VALUE_COLS[i]}{f['row']}"
            if i < len(vals):
                kind = "num" if f["kind"] == "integer" else "str"
                sx.set_cell(ref, vals[i], kind=kind)
            else:
                sx.set_cell(ref, "", kind="empty")  # clear leftover example
            ext_modified.append(ref)
    sx.drop_hyperlinks(ext_modified)  # e.g. the example Logo link on an overwritten/cleared cell
    edited[fn] = sx.serialize()

    # schema sheets: replace example data rows with the user's rows (header kept)
    _fill_table(tpl, edited, SHEET_DICTIONARIES, dictionaries,
                ["code", "name", "description"],
                {"A": "17", "B": "4", "C": "4"},
                bool_cols=set())
    _fill_table(tpl, edited, SHEET_FIELDS, fields,
                ["dictionary_code", "name", "label", "type", "constraints",
                 "description", "uri", "entity", "cohort_filter"],
                {"A": "17", "B": "8", "C": "8", "D": "8", "E": "14",
                 "F": "8", "G": "8", "H": None, "I": None},
                bool_cols={"entity", "cohort_filter"})
    _fill_table(tpl, edited, SHEET_LOOKUPS, lookups,
                ["lookup", "name", "description", "uri"],
                {"A": "11", "B": "19", "C": "19", "D": "19"},
                bool_cols=set())
    return edited


def is_caps_placeholder(value):
    """True for the template's ALL-CAPS placeholder cells (e.g. 'TITLE',
    'LIST KEYWORDS', 'IDENTETIFIER') — cleared when the user provides no value. Genuine
    defaults are mixed- or lower-case ('en', a URL, 'GEO/ArrayExpress') and are kept."""
    if value is None:
        return False
    s = str(value).strip()
    return s != "" and s == s.upper() and s != s.lower()


def _load_sheet(tpl, fn):
    with zipfile.ZipFile(tpl.path) as z:
        return SheetXML(z.read(fn).decode("utf-8"))


def _fill_table(tpl, edited, sheet_name, df, columns, col_styles, bool_cols):
    """Replace data rows (below header row 1) of a schema sheet with df's rows.

    If df is None the sheet is left untouched. If df is empty the example rows are
    cleared, leaving just the header."""
    if df is None:
        return
    fn = tpl._sheet_filename(sheet_name)
    sx = _load_sheet(tpl, fn)
    col_letters = [chr(ord("A") + i) for i in range(len(columns))]
    new_rows = []
    for _, r in df.iterrows():
        cells = {}
        for letter, colname in zip(col_letters, columns):
            val = r.get(colname, "")
            style = col_styles.get(letter)
            if colname in bool_cols:
                if str(val).strip() == "":
                    continue
                cells[letter] = ("bool", val, style)
            else:
                cells[letter] = ("auto", val, style)  # "" -> empty styled cell
        new_rows.append(cells)
    # materialize with correct row numbers
    header_row = 1
    built = []
    for i, cells in enumerate(new_rows):
        rownum = header_row + 1 + i
        rowcells = {}
        for letter, (kind, val, style) in cells.items():
            ref = f"{letter}{rownum}"
            rowcells[ref] = SheetXML.build_cell(ref, val, style=style, kind=kind)
        built.append(rowcells)
    sx.replace_data_rows(header_row, built)
    edited[fn] = sx.serialize()


# ===========================================================================
# Repackaging the zip (copy everything verbatim, swap edited worksheet XMLs)
# ===========================================================================
def write_zip(src_path, out_path, edited):
    if os.path.abspath(src_path) == os.path.abspath(out_path):
        sys.exit("refusing to overwrite the template; choose a different --out path.")
    tmp = out_path + ".part"
    with zipfile.ZipFile(src_path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in edited:
                data = edited[item.filename].encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, out_path)


# ===========================================================================
# CLI
# ===========================================================================
def main(argv=None):
    p = argparse.ArgumentParser(
        prog="addi",
        description="Fill the UK DRI FAIR-metadata workbook for an ADDI / AD "
                    "Workbench submission (never submits — writes files only).")
    sub = p.add_subparsers(dest="cmd", required=True)

    def add_template(sp):
        sp.add_argument("--template", default=DEFAULT_TEMPLATE,
                        help="template .xlsx (default: shipped V1.2 workbook)")

    sp = sub.add_parser("info", help="describe the template and its vocabularies")
    add_template(sp)
    sp.add_argument("--json", action="store_true", help="machine-readable dump")
    sp.set_defaults(func=cmd_info)

    sp = sub.add_parser("fill", help="write inputs into the template -> submission xlsx")
    add_template(sp)
    sp.add_argument("--config", help="JSON/YAML with catalogue/extendedcatalogue/"
                                     "settings/workspace_settings values")
    sp.add_argument("--metadata", help="metadata.tsv to seed the schema + some "
                                       "descriptive fields (lowest precedence)")
    sp.add_argument("--metadata-dict-code", default="sample_metadata",
                    help="dictionary code for the metadata-seeded table")
    sp.add_argument("--dictionaries", help="dictionaries table (TSV/CSV)")
    sp.add_argument("--fields", help="fields table (TSV/CSV)")
    sp.add_argument("--lookups", help="lookups table (TSV/CSV)")
    sp.add_argument("--out", default="UK_DRI_FAIR_Metadata_filled.xlsx",
                    help="output workbook path")
    sp.set_defaults(func=cmd_fill)

    sp = sub.add_parser("validate", help="check inputs or a filled workbook")
    add_template(sp)
    sp.add_argument("--workbook", help="a filled .xlsx to validate (instead of inputs)")
    sp.add_argument("--config", help="JSON/YAML config to validate")
    sp.add_argument("--dictionaries", help="dictionaries table (TSV/CSV)")
    sp.add_argument("--fields", help="fields table (TSV/CSV)")
    sp.add_argument("--lookups", help="lookups table (TSV/CSV)")
    sp.set_defaults(func=cmd_validate)

    args = p.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
